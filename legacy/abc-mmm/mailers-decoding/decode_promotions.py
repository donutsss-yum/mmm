import re
import warnings
from datetime import date

import pandas as pd

warnings.filterwarnings("ignore")

PROMO_FILES = {
    2024: "ABC Promotions 2024.xlsx",
    2025: "ABC Promotions 2025.xlsx",
    2026: "ABC Promotions 2026.xlsx",
}
MAILERS_FILE = "Mailers.xlsx"

# label (normalized) -> output field, for the metric rows inside a sub-block.
# Order of checks matters; resolved in norm_label() via longest/most-specific first.
METRIC_MAP = [
    ("av retail value w/ sale prices and", "avg_retail_value_w_sale_promo"),
    ("avg retail value w/ sale prices", "avg_retail_value_w_sale"),
    ("avg full retail value", "avg_full_retail_value"),
    ("full retail value", "full_retail_value"),
    ("retail value", "gross_rev"),  # 2022 label for what later years call "Gross Rev"
    ("markdown discount", "markdown_discount"),
    ("total promotional discount", "total_promotional_discount"),
    ("promo discount (this promo)", "promo_discount_this_promo"),
    ("promotional discount", "total_promotional_discount"),  # 2022/2023 single-line label
    ("gross rev", "gross_rev"),
    ("# of orders", "num_orders"),
    ("promotional cost per order w/ sale", "promo_cost_per_order_w_sale"),
    ("promotional cost per order w/ promo", "promo_cost_per_order_w_promo"),
    ("promo sales", "promo_sales_trans_total"),
    ("gross profit margin", "gross_profit_margin"),
    ("transactions", "transactions"),
    ("total discount", "total_discount"),
    ("aov", "aov"),
]

NUMERIC_FIELDS = [
    "full_retail_value", "markdown_discount", "total_promotional_discount",
    "promo_discount_this_promo", "gross_rev", "avg_full_retail_value",
    "avg_retail_value_w_sale", "avg_retail_value_w_sale_promo",
    "promo_cost_per_order_w_sale", "promo_cost_per_order_w_promo", "aov",
    "promo_sales_trans_total", "gross_profit_margin", "total_discount",
    "discount_value", "min_spend",
]
INT_FIELDS = ["num_orders", "transactions", "year", "duration_days"]
STRING_FIELDS = [
    "source_file", "category", "channel", "promotion_name", "description_raw",
    "discount_type", "discount_unit", "product_scope", "sizes",
]

COLUMN_ORDER = (
    ["source_file", "year", "category", "channel", "promotion_name",
     "start_date", "end_date", "duration_days",
     "discount_type", "discount_value", "discount_unit", "min_spend",
     "product_scope", "sizes", "description_raw"]
    + ["full_retail_value", "num_orders"]
    + ["gross_rev", "markdown_discount", "total_promotional_discount",
       "promo_discount_this_promo", "avg_full_retail_value",
       "avg_retail_value_w_sale", "avg_retail_value_w_sale_promo",
       "promo_cost_per_order_w_sale", "promo_cost_per_order_w_promo", "aov",
       "promo_sales_trans_total", "transactions", "gross_profit_margin",
       "total_discount"]
)

SECTION_LABELS = {"occ", "instore"}


def norm(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\xa0", " ").strip()
    if s.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", s)


def is_valid_promo_name(name):
    """Reject empty, 'Totals', and stray date/numeric header cells (sheet noise)."""
    n = norm(name)
    if not n or n.lower() == "totals":
        return False
    if re.fullmatch(r"[\d\s.,:/-]+", n):  # pure number / date-like
        return False
    if re.match(r"\d{4}-\d{2}-\d{2}", n):  # ISO datetime leak
        return False
    return True


def infer_category(name, desc, cat_map):
    """Summary-derived map first, then regex on name/description."""
    n = norm(name)
    if n in cat_map:
        return cat_map[n]
    blob = (n + " " + norm(desc)).lower()
    if "birthday" in blob:
        return "Birthday"
    if "free bottle" in blob or "get a free" in blob or "free wine" in blob:
        return "Free Bottles"
    if "vbc" in blob:
        return "VBCs"
    if "mailer" in blob:
        return "Mailer"
    if "sign up" in blob or "sign-up" in blob:
        return "New Sign Ups"
    return "Other"


def build_category_map(files):
    """promotion_name -> category, harvested from the (partial) Summary sheets."""
    cmap = {}
    for fname in files:
        if "Summary" not in pd.ExcelFile(fname).sheet_names:
            continue  # 2022 has no Summary sheet; categories come via regex
        df = pd.read_excel(fname, sheet_name="Summary", header=None)
        category = None
        for i in range(len(df)):
            row = df.iloc[i]
            c0 = norm(row.iloc[0])
            nonblank = [norm(x) for x in row.iloc[1:].tolist() if norm(x)]
            if c0 and not nonblank and c0.lower() not in SECTION_LABELS and not match_metric(c0):
                category = c0
            elif c0.lower() in SECTION_LABELS and category:
                for ci in range(1, df.shape[1]):
                    nm = norm(row.iloc[ci])
                    if is_valid_promo_name(nm):
                        cmap.setdefault(nm, category)
    return cmap


def match_metric(label):
    n = norm(label).lower()
    if not n:
        return None
    for key, field in METRIC_MAP:
        if n.startswith(key):
            return field
    return None


def parse_one_date(token):
    """'1/1', '12/26/2024', '4/04/2026' -> (month, day, year|None)."""
    token = token.strip().strip(",")
    if not token:
        return None
    parts = token.split("/")
    try:
        if len(parts) == 3:
            m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return (m, d, y)
        if len(parts) == 2:
            return (int(parts[0]), int(parts[1]), None)
    except ValueError:
        return None
    return None


def parse_date_range(raw, file_year):
    """Parse a 'Dates Active' cell into ISO (start, end) strings.

    Handles every observed variant: 'M/D-M/D' (no year), 'M/D-M/D YYYY'
    (trailing year), 'M/D/YYYY-M/D/YYYY', 'M/D-M/D/YYYY', cross-year ranges
    (12/12-1/4), and a single datetime cell (one-day event -> start==end).
    Year precedence: explicit slashed year > trailing year > file year;
    a one-sided year propagates to the other side; cross-year inferred
    when the end M/D precedes the start M/D.
    """
    s = norm(raw)
    if not s or s.lower() in ("nan", "totals"):
        return (None, None)

    iso = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)  # datetime cell leak
    if iso:
        try:
            d = date(int(iso.group(1)), int(iso.group(2)), int(iso.group(3)))
            return (d.isoformat(), d.isoformat())
        except ValueError:
            return (None, None)

    s = s.replace("–", "-").replace("—", "-")
    trailing_year = None
    m_ty = re.search(r"[-\s](\d{4})\s*$", s)
    if m_ty and "/" + m_ty.group(1) not in s:  # ' 2024' suffix, not M/D/2024
        trailing_year = int(m_ty.group(1))
        s = s[: m_ty.start()].strip()

    if "-" not in s:  # single date -> one-day event
        a = parse_one_date(s)
        if a is None:
            return (None, None)
        m, d, y = a
        y = y or trailing_year or file_year
        try:
            dt = date(y, m, d)
            return (dt.isoformat(), dt.isoformat())
        except ValueError:
            return (None, None)

    left, right = s.split("-", 1)
    a, b = parse_one_date(left), parse_one_date(right)
    if a is None or b is None:
        return (None, None)
    sm, sd, sy = a
    em, ed, ey = b
    if sy is None:
        sy = trailing_year
    if ey is None:
        ey = trailing_year

    if sy is None and ey is None:
        sy = ey = file_year
        if (em, ed) < (sm, sd):
            ey = sy + 1
    elif sy is None:
        sy = ey - 1 if (sm, sd) > (em, ed) else ey
    elif ey is None:
        ey = sy + 1 if (em, ed) < (sm, sd) else sy

    try:
        sdt, edt = date(sy, sm, sd), date(ey, em, ed)
    except ValueError:
        return (None, None)
    return (sdt.isoformat(), edt.isoformat())


def parse_description(text):
    """Best-effort structured parse; description_raw always retained upstream."""
    t = norm(text)
    out = {
        "discount_type": None, "discount_value": None, "discount_unit": None,
        "min_spend": None, "product_scope": None, "sizes": None,
    }
    if not t:
        return out
    low = t.lower()

    m_min = re.search(r"\$\s*(\d+(?:\.\d+)?)\s*\+", t)
    if m_min:
        out["min_spend"] = float(m_min.group(1))
    m_spend = re.search(r"spend\s*\$\s*(\d+(?:\.\d+)?)", low)
    if m_spend and out["min_spend"] is None:
        out["min_spend"] = float(m_spend.group(1))

    m_pct = re.search(r"(\d+(?:\.\d+)?)\s*%\s*off", low)
    m_dol = re.search(r"\$\s*(\d+(?:\.\d+)?)\s*off", low)
    is_free = ("free" in low) or ("get a free" in low)
    if m_pct:
        out["discount_type"] = "percent_off"
        out["discount_value"] = float(m_pct.group(1))
        out["discount_unit"] = "percent"
    elif m_dol:
        out["discount_type"] = "dollar_off"
        out["discount_value"] = float(m_dol.group(1))
        out["discount_unit"] = "USD"
    elif is_free:
        out["discount_type"] = "spend_get_free" if "spend" in low else "free_item"
        out["discount_unit"] = "free_item"

    sizes = re.findall(r"\d+(?:\.\d+)?\s*(?:mL|ML|ml|L|l)\b", t)
    if not sizes:
        m_sz = re.search(r"\d+\s*mL[^.;]*sizes", t)
        if m_sz:
            sizes = [m_sz.group(0)]
    if sizes:
        out["sizes"] = ", ".join(dict.fromkeys(s.strip() for s in sizes))

    scope = None
    m_scope = re.search(r"off\s+(?:\$\s*\d+\+?\s+)?(.*)", t, re.IGNORECASE)
    if m_scope:
        scope = m_scope.group(1)
    elif is_free:
        m_g = re.search(r"get\s+a?\s*(.*)", t, re.IGNORECASE)
        if m_g:
            scope = m_g.group(1)
    if scope:
        scope = re.split(r"\d+(?:\.\d+)?\s*(?:mL|ML|ml|L|l)\b", scope)[0]
        scope = re.sub(r"\bp(?:code|codes)\b.*$", "", scope, flags=re.IGNORECASE)
        scope = scope.strip(" ,.;-")
        out["product_scope"] = scope or None
    return out


def block_records(df, file_year, source_file):
    """Walk a sheet of stacked category -> (OCC|INSTORE) -> metric-row blocks."""
    records = []
    category = None
    n_rows, n_cols = df.shape

    r = 0
    while r < n_rows:
        row = df.iloc[r]
        c0 = norm(row.iloc[0])
        nonblank = [norm(x) for x in row.iloc[1:].tolist() if norm(x)]

        if c0 and not nonblank and c0.lower() not in SECTION_LABELS and not match_metric(c0):
            category = c0  # lone title cell => category header
            r += 1
            continue

        if c0.lower() in SECTION_LABELS:
            channel = "OCC" if c0.lower() == "occ" else "INSTORE"
            # promo columns: header cells, excluding trailing 'Totals'
            promo_cols = []
            for ci in range(1, n_cols):
                name = norm(row.iloc[ci])
                if not is_valid_promo_name(name):
                    continue
                promo_cols.append((ci, name))

            data = {ci: {"promotion_name": name} for ci, name in promo_cols}
            desc_row_idx = None
            r += 1
            while r < n_rows:
                rr = df.iloc[r]
                lbl = norm(rr.iloc[0])
                if lbl.lower() in SECTION_LABELS:
                    break
                lbl_nonblank = [norm(x) for x in rr.iloc[1:].tolist() if norm(x)]
                if lbl and not lbl_nonblank and lbl.lower() not in SECTION_LABELS and not match_metric(lbl):
                    break  # next category title

                low = lbl.lower()
                if low == "description":
                    desc_row_idx = r
                    for ci, _ in promo_cols:
                        data[ci]["description_raw"] = norm(rr.iloc[ci])
                elif low.startswith("dates active"):
                    for ci, _ in promo_cols:
                        data[ci]["_dates"] = norm(rr.iloc[ci])
                elif not lbl and desc_row_idx is not None and r == desc_row_idx + 1:
                    # unlabeled continuation row (Mailers): append to description
                    for ci, _ in promo_cols:
                        cont = norm(rr.iloc[ci])
                        if cont:
                            data[ci]["description_raw"] = (
                                data[ci].get("description_raw", "") + " " + cont
                            ).strip()
                else:
                    field = match_metric(lbl)
                    if field:
                        for ci, _ in promo_cols:
                            data[ci][field] = rr.iloc[ci]
                r += 1

            for ci, _ in promo_cols:
                rec = data[ci]
                start, end = parse_date_range(rec.pop("_dates", ""), file_year)
                desc = rec.get("description_raw", "")
                parsed = parse_description(desc)
                out = {
                    "source_file": source_file,
                    "year": file_year,
                    "category": category,
                    "channel": channel,
                    "promotion_name": rec.get("promotion_name"),
                    "start_date": start,
                    "end_date": end,
                    "description_raw": desc or None,
                }
                out.update(parsed)
                for _, fld in METRIC_MAP:
                    if fld in rec:
                        out[fld] = rec[fld]
                records.append(out)
            continue
        r += 1
    return records


def coerce_types(df):
    for c in NUMERIC_FIELDS:
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    for c in INT_FIELDS:
        if c in df:
            df[c] = pd.to_numeric(df[c], errors="coerce").round().astype("Int64")
    for c in STRING_FIELDS:
        if c in df:
            df[c] = df[c].apply(lambda v: None if (v is None or (isinstance(v, float) and pd.isna(v)) or norm(v) == "") else norm(v))
    for c in ("start_date", "end_date"):
        df[c] = pd.to_datetime(df[c], errors="coerce").dt.date
    sd = pd.to_datetime(df["start_date"], errors="coerce")
    ed = pd.to_datetime(df["end_date"], errors="coerce")
    df["duration_days"] = ((ed - sd).dt.days + 1).astype("Int64")
    # calendar year of the promo (from start_date) for time-series use;
    # fall back to the workbook year when the source has no date
    df["year"] = sd.dt.year.astype("Int64").fillna(
        pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    )
    return df


DEDUP_KEY = ["source_file", "channel", "promotion_name", "start_date",
             "end_date", "description_raw"]


def bq_type(col):
    if col in ("start_date", "end_date"):
        return "DATE"
    if col in INT_FIELDS:
        return "INT64"
    if col in NUMERIC_FIELDS:
        return "FLOAT64"
    return "STRING"


def write_bq_schema(df, schema_path, ddl_path):
    import json

    schema = [{"name": c, "type": bq_type(c), "mode": "NULLABLE"}
              for c in df.columns]
    with open(schema_path, "w") as fh:
        json.dump(schema, fh, indent=2)
    cols = ",\n".join(f"  {c} {bq_type(c)}" for c in df.columns)
    ddl = (
        "-- One row per promotion per channel (OCC / INSTORE).\n"
        "-- Load: bq load --source_format=CSV --skip_leading_rows=1 \\\n"
        "--   --null_marker='' dataset.abc_promotions \\\n"
        "--   ABC_promotions_clean.csv ABC_promotions_clean.bq_schema.json\n"
        f"CREATE TABLE IF NOT EXISTS dataset.abc_promotions (\n{cols}\n);\n"
    )
    with open(ddl_path, "w") as fh:
        fh.write(ddl)


def write_excel(df, path):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    xdf = df.copy()
    for c in ("start_date", "end_date"):
        xdf[c] = pd.to_datetime(xdf[c], errors="coerce")
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="YYYY-MM-DD") as xw:
        xdf.to_excel(xw, sheet_name="promotions", index=False)
        ws = xw.sheets["promotions"]
        head_fill = PatternFill("solid", fgColor="1F4E78")
        head_font = Font(name="Arial", bold=True, color="FFFFFF")
        for ci, col in enumerate(xdf.columns, 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            width = max(len(str(col)) + 2,
                        int(xdf[col].astype(str).str.len().quantile(0.92) or 8) + 2)
            ws.column_dimensions[get_column_letter(ci)].width = min(width, 46)
        money = {"full_retail_value", "gross_rev", "markdown_discount",
                 "total_promotional_discount", "promo_discount_this_promo",
                 "avg_full_retail_value", "avg_retail_value_w_sale",
                 "avg_retail_value_w_sale_promo", "promo_cost_per_order_w_sale",
                 "promo_cost_per_order_w_promo", "aov", "promo_sales_trans_total",
                 "total_discount", "min_spend", "discount_value"}
        body = Font(name="Arial")
        for ci, col in enumerate(xdf.columns, 1):
            for ri in range(2, ws.max_row + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.font = body
                if col in ("start_date", "end_date"):
                    cell.number_format = "YYYY-MM-DD"
                elif col in money:
                    cell.number_format = '#,##0.00;(#,##0.00);-'
                elif col in ("num_orders", "transactions", "duration_days", "year"):
                    cell.number_format = '0'
        ws.freeze_panes = "F2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(xdf.columns))}{ws.max_row}"


def main():
    cat_map = build_category_map(PROMO_FILES.values())

    all_recs = []
    for yr, fname in PROMO_FILES.items():
        xl = pd.ExcelFile(fname)
        n_before = len(all_recs)
        for sheet in xl.sheet_names:
            if sheet == "Summary":  # partial rollup; monthly sheets are complete
                continue
            df = pd.read_excel(fname, sheet_name=sheet, header=None)
            recs = block_records(df, yr, fname.replace(".xlsx", ""))
            for rec in recs:
                rec["category"] = infer_category(
                    rec.get("promotion_name"), rec.get("description_raw"), cat_map
                )
            all_recs.extend(recs)
        print(f"{fname}: {len(all_recs) - n_before} promo-channel rows "
              f"(from {len(xl.sheet_names) - 1} monthly sheets)")

    mdf = pd.read_excel(MAILERS_FILE, sheet_name="Sheet1", header=None)
    mrecs = block_records(mdf, None, "Mailers")
    for rrec in mrecs:
        rrec["category"] = "Mailer"
    all_recs.extend(mrecs)
    print(f"{MAILERS_FILE}: {len(mrecs)} promo-channel rows")

    df = pd.DataFrame(all_recs)
    df = df.dropna(subset=["promotion_name"]).reset_index(drop=True)
    before = len(df)
    df = df.drop_duplicates(subset=DEDUP_KEY, keep="first").reset_index(drop=True)
    print(f"Deduplicated multi-month repeats: {before} -> {len(df)} rows")
    for col in COLUMN_ORDER:
        if col not in df:
            df[col] = pd.NA
    df = df[COLUMN_ORDER]
    df = coerce_types(df)
    df = df.sort_values(
        ["start_date", "source_file", "category", "channel", "promotion_name"],
        na_position="last",
    ).reset_index(drop=True)

    df.to_csv("ABC_promotions_clean.csv", index=False, date_format="%Y-%m-%d")
    write_excel(df, "ABC_promotions_clean.xlsx")
    write_bq_schema(df, "ABC_promotions_clean.bq_schema.json",
                    "ABC_promotions_clean.bq.sql")

    sd = pd.to_datetime(df["start_date"], errors="coerce")
    ed = pd.to_datetime(df["end_date"], errors="coerce")
    print("\nTotal rows:", len(df))
    print("Date range:", sd.min().date(), "->", ed.max().date())
    print("Null start_date rows:", int(df["start_date"].isna().sum()))
    print("Rows w/ full_retail_value:", int(df["full_retail_value"].notna().sum()),
          "| Rows w/ num_orders:", int(df["num_orders"].notna().sum()))
    print("Outputs: ABC_promotions_clean.csv / .xlsx / .bq_schema.json / .bq.sql")
    return df


if __name__ == "__main__":
    main()
